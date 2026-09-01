"""IA Brasil — Export Service.

Serviço de geração de arquivos CSV, JSON, XLSX e PDF para exportação de dados abertos.
Conforme issue #19: exportação de dados abertos (CSV e JSON) por eixo e geral.
Conforme issue #1030: exportação em mais formatos (PDF, XLSX).

Estrutura do CSV exportado:
    eixo_codigo, eixo_nome, programa_nome, acao_codigo, acao_nome,
    status_atual, data_avaliacao, instituicoes, meta_count, source_ref
"""

from __future__ import annotations

import asyncio
import csv
import io
from typing import TYPE_CHECKING

from sqlalchemy import func, select
from sqlalchemy.orm import joinedload, selectinload

from src.core.json_encoder import dumps_with_encoder

if TYPE_CHECKING:
    from collections.abc import AsyncGenerator
    from typing import Any

    from sqlalchemy.ext.asyncio import AsyncSession

from src.core.db import (
    Acao,
    AcaoInstituicao,
    Eixo,
    Evento,
    Programa,
    get_session,
)

# ============================================================================
# Utilitários
# ============================================================================

# Cabeçalhos padrão para CSV de exportação de ações
CSV_ACOES_HEADERS: list[str] = [
    "eixo_codigo",
    "eixo_nome",
    "programa_nome",
    "acao_codigo",
    "acao_nome",
    "status_atual",
    "data_avaliacao",
    "instituicoes",
    "meta_count",
    "source_ref",
]


async def get_db() -> AsyncGenerator[AsyncSession, None]:
    """Dependency para obter sessão do banco."""
    async with get_session() as session:
        yield session


# ============================================================================
# Serviço de Exportação
# ============================================================================


class ExportService:
    """Serviço para geração de arquivos de exportação."""

    @staticmethod
    async def get_all_acoes_data(
        session: AsyncSession,
    ) -> list[dict[str, Any]]:
        """Obter dados de todas as ações para exportação.

        Args:
            session: Sessão assíncrona do banco de dados

        Returns:
            Lista de dicionários com dados das ações formatados para exportação

        Note:
            Sempre usa nome real da ação (acao.nome) via join, nunca formato
            simplificado como "Ação {id}"
        """
        # Obter todas as ações com seus relacionamentos usando eager loading
        # para evitar N+1 queries conforme issue #110
        result = await session.execute(
            select(Acao)
            .join(Programa, Acao.programa_id == Programa.id)
            .join(Eixo, Programa.eixo_id == Eixo.id)
            .options(
                joinedload(Acao.programa).joinedload(Programa.eixo),
                selectinload(Acao.avaliacoes),
                selectinload(Acao.instituicoes).selectinload(AcaoInstituicao.instituicao),
                selectinload(Acao.metas),
            )
            .order_by(Eixo.numero, Programa.nome, Acao.nome)
        )
        acoes = result.scalars().all()

        export_data = []
        for acao in acoes:
            # Obter avaliação mais recente (já carregada via eager loading)
            # Ordenar manualmente as avaliações já carregadas
            avaliacao = None
            if acao.avaliacoes:
                avaliacao = sorted(
                    acao.avaliacoes, key=lambda x: (x.data_avaliacao, x.versao), reverse=True
                )[0]

            # Obter instituições (já carregadas via eager loading)
            instituicoes = [f"{ai.instituicao.sigla} ({ai.papel})" for ai in acao.instituicoes]

            # Obter contagem de metas (já carregadas via eager loading)
            meta_count = len(acao.metas)

            export_data.append(
                {
                    "eixo_codigo": acao.programa.eixo.numero,
                    "eixo_nome": acao.programa.eixo.nome,
                    "programa_nome": acao.programa.nome,
                    "acao_codigo": acao.codigo_oficial or f"A{acao.id[:8]}",
                    "acao_nome": acao.nome,
                    "status_atual": (
                        avaliacao.status_avaliado.value if avaliacao else acao.status.value
                    ),
                    "data_avaliacao": (avaliacao.data_avaliacao.isoformat() if avaliacao else None),
                    "instituicoes": "; ".join(instituicoes) if instituicoes else "",
                    "meta_count": meta_count,
                    "source_ref": f"pbia:{acao.id}",
                }
            )

        return export_data

    @staticmethod
    async def get_acoes_by_eixo_data(
        session: AsyncSession,
        eixo_id: str,
    ) -> list[dict[str, Any]]:
        """Obter dados de ações filtradas por eixo para exportação.

        Args:
            session: Sessão assíncrona do banco de dados
            eixo_id: ID do eixo para filtrar

        Returns:
            Lista de dicionários com dados das ações formatados para exportação

        Note:
            Sempre usa nome real da ação (acao.nome) via join, nunca formato
            simplificado como "Ação {id}"
        """
        # Obter ações do eixo específico com eager loading para evitar N+1 queries
        result = await session.execute(
            select(Acao)
            .join(Programa, Acao.programa_id == Programa.id)
            .join(Eixo, Programa.eixo_id == Eixo.id)
            .where(Eixo.id == eixo_id)
            .options(
                joinedload(Acao.programa).joinedload(Programa.eixo),
                selectinload(Acao.avaliacoes),
                selectinload(Acao.instituicoes).selectinload(AcaoInstituicao.instituicao),
                selectinload(Acao.metas),
            )
            .order_by(Programa.nome, Acao.nome)
        )
        acoes = result.scalars().all()

        export_data = []
        for acao in acoes:
            # Obter avaliação mais recente (já carregada via eager loading)
            # Ordenar manualmente as avaliações já carregadas
            avaliacao = None
            if acao.avaliacoes:
                avaliacao = sorted(
                    acao.avaliacoes, key=lambda x: (x.data_avaliacao, x.versao), reverse=True
                )[0]

            # Obter instituições (já carregadas via eager loading)
            instituicoes = [f"{ai.instituicao.sigla} ({ai.papel})" for ai in acao.instituicoes]

            # Obter contagem de metas (já carregadas via eager loading)
            meta_count = len(acao.metas)

            export_data.append(
                {
                    "eixo_codigo": acao.programa.eixo.numero,
                    "eixo_nome": acao.programa.eixo.nome,
                    "programa_nome": acao.programa.nome,
                    "acao_codigo": acao.codigo_oficial or f"A{acao.id[:8]}",
                    "acao_nome": acao.nome,
                    "status_atual": (
                        avaliacao.status_avaliado.value if avaliacao else acao.status.value
                    ),
                    "data_avaliacao": (avaliacao.data_avaliacao.isoformat() if avaliacao else None),
                    "instituicoes": "; ".join(instituicoes) if instituicoes else "",
                    "meta_count": meta_count,
                    "source_ref": f"pbia:{acao.id}",
                }
            )

        return export_data

    @staticmethod
    async def get_all_eventos_data(
        session: AsyncSession,
    ) -> list[dict[str, Any]]:
        """Obter dados de todos os eventos para exportação.

        Args:
            session: Sessão assíncrona do banco de dados

        Returns:
            Lista de dicionários com dados dos eventos formatados para exportação

        Note:
            Sempre usa nome real da ação (acao.nome) via join, nunca formato
            simplificado como "Ação {id}"
        """
        # Obter todos os eventos com seus relacionamentos usando eager loading
        # selectinload(Evento.acao) evita N+1 queries ao carregar ações associadas
        # em uma única consulta, otimizado para datasets grandes conforme issue #55
        result = await session.execute(
            select(Evento)
            .options(selectinload(Evento.acao))
            .order_by(Evento.data_evento.desc(), Evento.criado_em.desc())
        )
        eventos = result.scalars().all()

        export_data = []
        for evento in eventos:
            # Obter nome da ação com fallback seguro
            # Usa nome real da ação via eager loading, nunca formato simplificado
            acao_nome = (
                evento.acao.nome
                if evento.acao
                else f"Ação não encontrada (ID: {evento.acao_id}) — possível inconsistência"
            )

            export_data.append(
                {
                    "evento_id": evento.id,
                    "acao_id": evento.acao_id,
                    "acao_nome": acao_nome,
                    "evento_tipo": evento.tipo.value,
                    "evento_descricao": evento.descricao,
                    "data_evento": evento.data_evento.isoformat(),
                    "criado_em": evento.criado_em.isoformat(),
                    "fonte_url": evento.fonte_url,
                    "referencia_id": evento.referencia_id,
                    "referencia_tipo": evento.referencia_tipo,
                }
            )

        return export_data

    @staticmethod
    def generate_csv(data: list[dict[str, Any]]) -> str:
        """Gerar conteúdo CSV a partir dos dados.

        Args:
            data: Lista de dicionários com dados para exportação

        Returns:
            String com conteúdo CSV incluindo BOM UTF-8 para compatibilidade com Excel
        """
        # Criar buffer com BOM UTF-8 para Excel
        output = io.StringIO()
        output.write("\ufeff")  # BOM UTF-8

        # Determinar cabeçalhos: usar do primeiro item se houver dados, caso contrário usar padrão
        fieldnames = list(data[0].keys()) if data else CSV_ACOES_HEADERS

        writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=",")
        writer.writeheader()

        if data:
            writer.writerows(data)

        return output.getvalue()

    @staticmethod
    def generate_json(data: list[dict[str, Any]]) -> str:
        """Gerar conteúdo JSON a partir dos dados.

        Args:
            data: Lista de dicionários com dados para exportação

        Returns:
            String com conteúdo JSON formatado
        """
        return dumps_with_encoder(data, ensure_ascii=False, indent=2)

    @staticmethod
    def generate_xlsx(data: list[dict[str, Any]], sheet_name: str = "Ações") -> bytes:
        """Gerar conteúdo XLSX a partir dos dados.

        Args:
            data: Lista de dicionários com dados para exportação
            sheet_name: Nome da planilha

        Returns:
            Bytes do arquivo XLSX
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

        # Cabeçalhos
        headers = list(data[0].keys())
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Dados
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        # Auto-fit column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                max_length + 2, 50
            )

        # Congelar primeira linha
        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def generate_pdf(data: list[dict[str, Any]], title: str = "Ações do PBIA") -> bytes:
        """Gerar conteúdo PDF a partir dos dados.

        Args:
            data: Lista de dicionários com dados para exportação
            title: Título do documento PDF

        Returns:
            Bytes do arquivo PDF
        """
        from fpdf import FPDF

        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 8, f"Total: {len(data)} registros", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(5)

        if not data:
            pdf.set_font("Helvetica", "", 12)
            pdf.cell(0, 10, "Nenhum dado disponível", new_x="LMARGIN", new_y="NEXT", align="C")
            buffer = io.BytesIO()
            pdf.output(buffer)
            return buffer.getvalue()

        # Cabeçalhos
        headers = list(data[0].keys())
        col_width = (297 - 20) / len(headers)  # A4 landscape width minus margins

        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(31, 78, 121)
        pdf.set_text_color(255, 255, 255)
        for header in headers:
            pdf.cell(col_width, 7, header[:30], border=1, fill=True, align="C")
        pdf.ln()

        # Dados
        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(0, 0, 0)
        for row_idx, row_data in enumerate(data):
            if row_idx % 2 == 0:
                pdf.set_fill_color(240, 240, 240)
            else:
                pdf.set_fill_color(255, 255, 255)
            for header in headers:
                value = str(row_data.get(header, ""))[:30]
                pdf.cell(col_width, 6, value, border=1, fill=True, align="L")
            pdf.ln()

            # Nova página a cada 40 linhas
            if (row_idx + 1) % 40 == 0 and row_idx + 1 < len(data):
                pdf.add_page()
                pdf.set_font("Helvetica", "B", 8)
                pdf.set_fill_color(31, 78, 121)
                pdf.set_text_color(255, 255, 255)
                for header in headers:
                    pdf.cell(col_width, 7, header[:30], border=1, fill=True, align="C")
                pdf.ln()
                pdf.set_font("Helvetica", "", 7)
                pdf.set_text_color(0, 0, 0)

        buffer = io.BytesIO()
        pdf.output(buffer)
        return buffer.getvalue()

    @staticmethod
    async def generate_xlsx_async(data: list[dict[str, Any]], sheet_name: str = "Ações") -> bytes:
        """Gera XLSX em thread separada para não bloquear o event loop.

        Args:
            data: Lista de dicionários com dados para exportação
            sheet_name: Nome da planilha

        Returns:
            Bytes do arquivo XLSX
        """
        return await asyncio.to_thread(ExportService.generate_xlsx, data, sheet_name)

    @staticmethod
    async def generate_pdf_async(data: list[dict[str, Any]], title: str = "Ações do PBIA") -> bytes:
        """Gera PDF em thread separada para não bloquear o event loop.

        Args:
            data: Lista de dicionários com dados para exportação
            title: Título do documento PDF

        Returns:
            Bytes do arquivo PDF
        """
        return await asyncio.to_thread(ExportService.generate_pdf, data, title)

    @staticmethod
    async def count_acoes(session: AsyncSession) -> int:
        """Conta o total de ações (usado no datapackage.json do export).

        Args:
            session: Sessão assíncrona do banco de dados

        Returns:
            Número total de ações no banco
        """
        result = await session.execute(select(func.count(Acao.id)))
        return int(result.scalar_one())

    @staticmethod
    async def get_eixo_name(session: AsyncSession, eixo_id: str) -> str:
        """Obter nome do eixo para uso em nomes de arquivos."""
        result = await session.execute(select(Eixo.nome).where(Eixo.id == eixo_id))
        return result.scalar_one_or_none() or "desconhecido"

    @staticmethod
    async def get_feed_data(session: AsyncSession) -> list[dict[str, Any]]:
        """Obter dados do feed para exportação.

        Args:
            session: Sessão assíncrona do banco de dados

        Returns:
            Lista de dicionários com dados das atividades formatados para exportação
        """
        from src.core.db import Acao, AuditLog, Evento

        # Obter eventos
        eventos_result = await session.execute(select(Evento).order_by(Evento.data_evento.desc()))
        eventos = eventos_result.scalars().all()

        # Obter mudanças de status
        changes_result = await session.execute(
            select(AuditLog).order_by(AuditLog.data_criacao.desc())
        )
        changes = changes_result.scalars().all()

        # Obter mapeamento de ações (mesmo padrão de get_timeline_data)
        acoes_result = await session.execute(select(Acao))
        acoes_map = {acao.id: acao.nome for acao in acoes_result.scalars().all()}

        export_data = []

        for evento in eventos:
            acao_id = evento.acao_id if evento.acao_id else ""
            acao_nome = acoes_map.get(acao_id, "Ação desconhecida")
            export_data.append(
                {
                    "id": evento.id,
                    "tipo": "evento",
                    "acao_id": evento.acao_id,
                    "acao_nome": acao_nome,
                    "titulo": f"[{evento.tipo}] {acao_nome}",
                    "descricao": evento.descricao,
                    "data": evento.data_evento.isoformat(),
                    "fonte_url": evento.fonte_url,
                }
            )

        for change in changes:
            status_anterior = change.status_anterior.value if change.status_anterior else None
            status_novo = change.status_novo.value
            change_acao_id = change.acao_id if change.acao_id else ""
            acao_nome = acoes_map.get(change_acao_id, "Ação desconhecida")
            export_data.append(
                {
                    "id": change.id,
                    "tipo": "status_change",
                    "acao_id": change.acao_id,
                    "acao_nome": acao_nome,
                    "titulo": f"Mudança de status: {status_anterior} → {status_novo}",
                    "descricao": change.justificativa,
                    "data": change.data_criacao.isoformat(),
                    "status_anterior": status_anterior,
                    "status_novo": status_novo,
                }
            )

        return export_data

    @staticmethod
    async def get_timeline_data(session: AsyncSession) -> list[dict[str, Any]]:
        """Obter dados do timeline para exportação.

        Args:
            session: Sessão assíncrona do banco de dados

        Returns:
            Lista de dicionários com dados do timeline formatados para exportação
        """
        from src.core.db import Acao, AuditLog, Evento

        # Obter eventos
        eventos_result = await session.execute(select(Evento).order_by(Evento.data_evento.desc()))
        eventos = eventos_result.scalars().all()

        # Obter mudanças de status
        changes_result = await session.execute(
            select(AuditLog).order_by(AuditLog.data_criacao.desc())
        )
        changes = changes_result.scalars().all()

        # Obter mapeamento de ações
        acoes_result = await session.execute(select(Acao))
        acoes_map = {acao.id: acao.nome for acao in acoes_result.scalars().all()}

        export_data = []

        for evento in eventos:
            evento_acao_id = evento.acao_id if evento.acao_id else ""
            export_data.append(
                {
                    "id": evento.id,
                    "tipo": "evento",
                    "acao_id": evento.acao_id,
                    "acao_nome": acoes_map.get(evento_acao_id, "Ação desconhecida"),
                    "evento_tipo": evento.tipo,
                    "descricao": evento.descricao,
                    "data": evento.data_evento.isoformat(),
                    "fonte_url": evento.fonte_url,
                }
            )

        for change in changes:
            status_anterior = change.status_anterior.value if change.status_anterior else None
            status_novo = change.status_novo.value
            change_acao_id = change.acao_id if change.acao_id else ""
            export_data.append(
                {
                    "id": change.id,
                    "tipo": "status_change",
                    "acao_id": change.acao_id,
                    "acao_nome": acoes_map.get(change_acao_id, "Ação desconhecida"),
                    "status_anterior": status_anterior,
                    "status_novo": status_novo,
                    "justificativa": change.justificativa,
                    "data": change.data_criacao.isoformat(),
                }
            )

        return export_data

    @staticmethod
    async def get_dashboard_data(session: AsyncSession) -> list[dict[str, Any]]:
        """Obter dados agregados do dashboard para exportação.

        Args:
            session: Sessão assíncrona do banco de dados

        Returns:
            Lista de dicionários com dados agregados formatados para exportação
        """
        from src.core.db import Acao, Eixo, Indicador, Meta, Programa, Recurso

        # Obter todas as ações
        acoes_result = await session.execute(select(Acao))
        acoes = acoes_result.scalars().all()

        # Obter programas e eixos
        programas_result = await session.execute(select(Programa).order_by(Programa.nome))
        programas = programas_result.scalars().all()

        eixos_result = await session.execute(select(Eixo))
        eixos = eixos_result.scalars().all()

        # Obter metas e indicadores
        metas_result = await session.execute(select(Meta))
        metas = metas_result.scalars().all()

        indicadores_result = await session.execute(select(Indicador))
        indicadores = indicadores_result.scalars().all()

        # Obter recursos
        recursos_result = await session.execute(select(Recurso))
        recursos = recursos_result.scalars().all()

        # Calcular estatísticas
        total_acoes = len(acoes)
        total_programas = len(programas)
        total_eixos = len(eixos)
        total_metas = len(metas)
        total_indicadores = len(indicadores)

        # Calcular recursos totais
        recursos_totais = sum(
            r.valor_previsto if r.valor_previsto is not None else 0 for r in recursos
        )

        # Calcular progresso por status
        status_counts: dict[str, int] = {}
        for acao in acoes:
            status = acao.status.value
            status_counts[status] = status_counts.get(status, 0) + 1

        # Format status summary
        status_summary = [
            {
                "status": status,
                "count": count,
                "percentage": round((count / total_acoes * 100) if total_acoes > 0 else 0, 1),
            }
            for status, count in status_counts.items()
        ]

        return [
            {
                "metric": "total_acoes",
                "value": total_acoes,
                "description": "Total de ações no PBIA",
            },
            {
                "metric": "total_programas",
                "value": total_programas,
                "description": "Total de programas",
            },
            {"metric": "total_eixos", "value": total_eixos, "description": "Total de eixos"},
            {"metric": "total_metas", "value": total_metas, "description": "Total de metas"},
            {
                "metric": "total_indicadores",
                "value": total_indicadores,
                "description": "Total de indicadores",
            },
            {
                "metric": "recursos_totais",
                "value": recursos_totais,
                "description": "Valor total previsto (R$)",
                "unidade": "R$",
            },
            {
                "metric": "status_distribution",
                "value": status_summary,
                "description": "Distribuição de ações por status",
            },
        ]


# ============================================================================
# Cache Management
# ============================================================================


class ExportCache:
    """Gerenciamento de cache para exportações."""

    # Cache simples em memória (em produção, usar Redis ou similar)
    _cache: dict[str, tuple[str, float]] = {}
    _cache_bytes: dict[str, tuple[bytes, float]] = {}

    @classmethod
    def get_cached(cls, cache_key: str) -> str | None:
        """Obter conteúdo cacheado se ainda válido."""
        import time

        cached = cls._cache.get(cache_key)
        if cached is None:
            return None

        content, timestamp = cached
        # Cache válido por 1 hora (3600 segundos)
        if time.time() - timestamp < 3600:
            return content

        return None

    @classmethod
    def set_cached(cls, cache_key: str, content: str) -> None:
        """Armazenar conteúdo no cache."""
        import time

        cls._cache[cache_key] = (content, time.time())

    @classmethod
    def get_cached_bytes(cls, cache_key: str) -> bytes | None:
        """Obter conteúdo bytes cacheado se ainda válido."""
        import time

        cached = cls._cache_bytes.get(cache_key)
        if cached is None:
            return None

        content, timestamp = cached
        if time.time() - timestamp < 3600:
            return content

        return None

    @classmethod
    def set_cached_bytes(cls, cache_key: str, content: bytes) -> None:
        """Armazenar conteúdo bytes no cache."""
        import time

        cls._cache_bytes[cache_key] = (content, time.time())

    @classmethod
    def clear_cache(cls) -> None:
        """Limpar todo o cache."""
        cls._cache.clear()
        cls._cache_bytes.clear()
